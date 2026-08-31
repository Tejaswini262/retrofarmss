from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, Cookie, Query
from fastapi.responses import JSONResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any
import os
import io
import uuid
import hmac
import hashlib
import logging
import bcrypt
import httpx
import razorpay
from pathlib import Path
from datetime import datetime, timezone, timedelta, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

MONGO_URL = os.environ['MONGO_URL']
DB_NAME = os.environ['DB_NAME']
RZP_KEY_ID = os.environ['RAZORPAY_KEY_ID']
RZP_KEY_SECRET = os.environ['RAZORPAY_KEY_SECRET']

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

rzp = razorpay.Client(auth=(RZP_KEY_ID, RZP_KEY_SECRET))

app = FastAPI(title="Retro Farms API")
api = APIRouter(prefix="/api")

# ==================== SEED DATA ====================

SEED_PRODUCTS = [
    {
        'slug': 'country-eggs', 'name': 'Country Eggs (Free Range)', 'category': 'eggs',
        'image': 'https://images.unsplash.com/photo-1498654077810-12c21d4d6dc3?auto=format&fit=crop&w=1000&q=80',
        'from_price': 180, 'description': 'Farm-fresh brown eggs from free-roaming country hens. Naturally rich yolks, no antibiotics, no cages.',
        'variants': [
            {'id': 'dozen', 'label': '1 Dozen (12 eggs)', 'price': 180, 'stock': 194},
            {'id': 'tray', 'label': '1 Tray (30 eggs)', 'price': 420, 'stock': 118},
        ],
    },
    {
        'slug': 'country-chicken', 'name': 'Country Chicken (Live Weight)', 'category': 'chicken',
        'image': 'https://images.unsplash.com/photo-1535275226173-7ee8b465f0c1?auto=format&fit=crop&w=1000&q=80',
        'from_price': 340, 'description': 'Free-range country chicken raised on grains and greens. Sold by live weight, cleaned & delivered same day.',
        'variants': [
            {'id': '1kg', 'label': '1 kg (approx.)', 'price': 340, 'stock': 22},
            {'id': '2kg', 'label': '2 kg (approx.)', 'price': 660, 'stock': 14},
        ],
    },
    {
        'slug': 'alphonso-mango', 'name': 'Alphonso Mango', 'category': 'fruits',
        'image': 'https://images.unsplash.com/photo-1553279768-865429fa0078?auto=format&fit=crop&w=1000&q=80',
        'from_price': 220, 'description': 'Sun-ripened Alphonso mangoes from our orchard. Sweet, aromatic and naturally grown.',
        'variants': [
            {'id': '1kg', 'label': '1 kg', 'price': 220, 'stock': 46},
            {'id': '3kg', 'label': '3 kg box', 'price': 620, 'stock': 20},
        ],
    },
    {
        'slug': 'guava', 'name': 'Guava', 'category': 'fruits',
        'image': 'https://images.unsplash.com/photo-1536511132770-e5058c7e8c46?auto=format&fit=crop&w=1000&q=80',
        'from_price': 90, 'description': 'Crisp, farm-picked guavas. Pesticide-free and packed with vitamin C.',
        'variants': [
            {'id': '1kg', 'label': '1 kg', 'price': 90, 'stock': 60},
            {'id': '2kg', 'label': '2 kg', 'price': 170, 'stock': 30},
        ],
    },
    {
        'slug': 'lemon', 'name': 'Lemon', 'category': 'fruits',
        'image': 'https://images.unsplash.com/photo-1590502593747-42a996133562?auto=format&fit=crop&w=1000&q=80',
        'from_price': 60, 'description': 'Juicy country lemons — thin skin, plenty of juice, chemical-free.',
        'variants': [
            {'id': '500g', 'label': '500 g', 'price': 60, 'stock': 80},
            {'id': '1kg', 'label': '1 kg', 'price': 110, 'stock': 55},
        ],
    },
    {
        'slug': 'sapota', 'name': 'Sapota (Chiku)', 'category': 'fruits',
        'image': 'https://images.unsplash.com/photo-1610970881699-44a5587cabec?auto=format&fit=crop&w=1000&q=80',
        'from_price': 110, 'description': 'Naturally-ripened sapota — soft, sweet and grown on our farm.',
        'variants': [{'id': '1kg', 'label': '1 kg', 'price': 110, 'stock': 40}],
    },
    {
        'slug': 'papaya', 'name': 'Papaya', 'category': 'fruits',
        'image': 'https://images.unsplash.com/photo-1617112848923-cc2234396a8d?auto=format&fit=crop&w=1000&q=80',
        'from_price': 80, 'description': 'Tree-ripened papayas, orange fleshed and full of flavor.',
        'variants': [{'id': 'each', 'label': '1 piece (approx 1kg)', 'price': 80, 'stock': 25}],
    },
    {
        'slug': 'moringa', 'name': 'Moringa (Drumsticks)', 'category': 'vegetables',
        'image': 'https://images.unsplash.com/photo-1666904854830-c5c0e7b6f6f1?auto=format&fit=crop&w=1000&q=80',
        'from_price': 45, 'description': 'Fresh drumsticks from farm moringa trees. Ideal for sambar and curries.',
        'variants': [{'id': '250g', 'label': '250 g bunch', 'price': 45, 'stock': 70}],
    },
    {
        'slug': 'bottle-gourd', 'name': 'Bottle Gourd', 'category': 'vegetables',
        'image': 'https://images.unsplash.com/photo-1615485500704-8e990f9900f7?auto=format&fit=crop&w=1000&q=80',
        'from_price': 50, 'description': 'Tender bottle gourds, hand-picked and pesticide-free.',
        'variants': [{'id': 'each', 'label': '1 piece', 'price': 50, 'stock': 34}],
    },
    {
        'slug': 'tomatoes', 'name': 'Tomatoes', 'category': 'vegetables',
        'image': 'https://images.unsplash.com/photo-1592924357228-91a4daadcfea?auto=format&fit=crop&w=1000&q=80',
        'from_price': 55, 'description': 'Vine-ripened country tomatoes bursting with flavor.',
        'variants': [{'id': '1kg', 'label': '1 kg', 'price': 55, 'stock': 90}],
    },
    {
        'slug': 'green-chilli', 'name': 'Green Chilli', 'category': 'vegetables',
        'image': 'https://images.unsplash.com/photo-1526346093744-3d4b6ee7f2f5?auto=format&fit=crop&w=1000&q=80',
        'from_price': 30, 'description': 'Farm-grown green chillies with a mild, aromatic heat.',
        'variants': [{'id': '250g', 'label': '250 g', 'price': 30, 'stock': 65}],
    },
]

# ==================== MODELS ====================

class Variant(BaseModel):
    id: str
    label: str
    price: int
    stock: int

class Product(BaseModel):
    slug: str
    name: str
    category: str
    image: str
    from_price: int
    description: str
    variants: List[Variant]

class CartItem(BaseModel):
    slug: str
    variant_id: str
    qty: int
    options: Optional[Dict[str, Any]] = None  # For chicken: bird_type, delivery_date, instructions, piece_size

class Address(BaseModel):
    full_name: str
    phone: str
    line1: str
    line2: Optional[str] = ''
    city: str
    pincode: str
    landmark: Optional[str] = ''
    lat: Optional[float] = None
    lng: Optional[float] = None

class OrderCreatePayload(BaseModel):
    items: List[CartItem]
    address: Address
    payment_method: str  # 'razorpay' or 'cod'

class PaymentVerifyPayload(BaseModel):
    order_id: str  # our order id
    razorpay_order_id: str
    razorpay_payment_id: str
    razorpay_signature: str

class UpdateOrderPayload(BaseModel):
    status: Optional[str] = None
    assigned_staff_id: Optional[str] = None

class UpdateStockPayload(BaseModel):
    stock: int

class AdminLoginPayload(BaseModel):
    email: str
    password: str

class StaffCreatePayload(BaseModel):
    name: str
    email: str
    phone: Optional[str] = ''
    password: str
    role: str = 'staff'  # 'staff' or 'admin'

class OfflineOrderItem(BaseModel):
    slug: str
    variant_id: str
    qty: int
    options: Optional[Dict[str, Any]] = None

class OfflineOrderPayload(BaseModel):
    customer_name: str
    customer_phone: str
    customer_email: Optional[str] = ''
    items: List[OfflineOrderItem]
    address: Optional[Dict[str, Any]] = None
    payment_method: str = 'offline'  # offline / cash / upi / bank / not_paid / cod
    payment_status: str = 'Paid'
    notes: Optional[str] = ''
    status: str = 'Placed'
    total_override: Optional[int] = None
    save_customer_address: bool = True  # save address to customer profile for reuse

class CategoryPayload(BaseModel):
    id: str
    label: str

class CategoryUpdatePayload(BaseModel):
    label: Optional[str] = None
    order: Optional[int] = None

class CustomerUpdatePayload(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[Dict[str, Any]] = None

class FeedbackCreatePayload(BaseModel):
    order_id: str
    rating: int = Field(ge=1, le=5)
    comment: Optional[str] = ''

class FeedbackEditPayload(BaseModel):
    rating: Optional[int] = Field(default=None, ge=1, le=5)
    comment: Optional[str] = None

class FeedbackModeratePayload(BaseModel):
    status: Optional[str] = None  # 'active' | 'flagged' | 'hidden'
    admin_response: Optional[str] = None

class VariantInput(BaseModel):
    id: str
    label: str
    price: int
    stock: int = 0

class ProductCreatePayload(BaseModel):
    slug: str
    name: str
    category: str
    image: str
    from_price: int
    description: str
    variants: List[VariantInput]

class ProductUpdatePayload(BaseModel):
    name: Optional[str] = None
    category: Optional[str] = None
    image: Optional[str] = None
    from_price: Optional[int] = None
    description: Optional[str] = None

class VariantUpdatePayload(BaseModel):
    label: Optional[str] = None
    price: Optional[int] = None
    stock: Optional[int] = None

class FarmerCreatePayload(BaseModel):
    name: str
    creds: str = ''
    role: str = ''
    photo: str = ''
    order: int = 0

class UpdateMePayload(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    current_password: Optional[str] = None
    new_password: Optional[str] = None

class FarmerUpdatePayload(BaseModel):
    name: Optional[str] = None
    creds: Optional[str] = None
    role: Optional[str] = None
    photo: Optional[str] = None
    order: Optional[int] = None

# ==================== HELPERS ====================

def now_utc():
    return datetime.now(timezone.utc)

def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()

def check_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False

async def get_current_user(request: Request) -> Optional[Dict[str, Any]]:
    token = request.cookies.get('session_token')
    if not token:
        auth = request.headers.get('Authorization', '')
        if auth.startswith('Bearer '):
            token = auth[7:]
    if not token:
        return None
    session = await db.sessions.find_one({'session_token': token}, {'_id': 0})
    if not session:
        return None
    exp = session.get('expires_at')
    if isinstance(exp, str):
        exp = datetime.fromisoformat(exp)
    if exp and exp.tzinfo is None:
        exp = exp.replace(tzinfo=timezone.utc)
    if exp and exp < now_utc():
        return None
    user = await db.users.find_one({'user_id': session['user_id']}, {'_id': 0})
    return user

async def require_user(request: Request):
    u = await get_current_user(request)
    if not u:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return u

async def require_admin_or_staff(request: Request):
    u = await require_user(request)
    if u.get('role') not in ('admin', 'staff'):
        raise HTTPException(status_code=403, detail="Admin/staff only")
    return u

async def require_admin(request: Request):
    u = await require_user(request)
    if u.get('role') != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    return u

def user_public(u: Dict[str, Any]) -> Dict[str, Any]:
    return {
        'user_id': u['user_id'],
        'email': u['email'],
        'name': u.get('name', ''),
        'picture': u.get('picture', ''),
        'role': u.get('role', 'customer'),
        'phone': u.get('phone', ''),
    }

# ==================== STARTUP ====================

@app.on_event("startup")
async def startup():
    # DB indexes for performance
    try:
        await db.users.create_index('email', unique=True)
        await db.users.create_index('phone')
        await db.users.create_index('role')
        await db.users.create_index('user_id', unique=True)
        await db.orders.create_index('order_id', unique=True)
        await db.orders.create_index('user_id')
        await db.orders.create_index([('created_at', -1)])
        await db.orders.create_index('status')
        await db.orders.create_index('payment_status')
        await db.sessions.create_index('session_token', unique=True)
        await db.sessions.create_index('expires_at')
        await db.products.create_index('slug', unique=True)
        await db.products.create_index('category')
        await db.farmers.create_index('farmer_id', unique=True)
        await db.categories.create_index('id', unique=True)
        await db.feedback.create_index('order_id', unique=True)
        await db.feedback.create_index('user_id')
        await db.feedback.create_index([('created_at', -1)])
        await db.feedback.create_index('status')
    except Exception as e:
        logging.warning(f"Index creation warning: {e}")

    # Seed products if empty
    if await db.products.count_documents({}) == 0:
        await db.products.insert_many([{**p} for p in SEED_PRODUCTS])
        logging.info("Seeded products")

    # Seed categories if empty
    if await db.categories.count_documents({}) == 0:
        default_cats = [
            {'id': 'eggs', 'label': 'Country Eggs', 'order': 1},
            {'id': 'chicken', 'label': 'Country Chicken', 'order': 2},
            {'id': 'fruits', 'label': 'Farm Fruits', 'order': 3},
            {'id': 'vegetables', 'label': 'Fresh Vegetables', 'order': 4},
        ]
        await db.categories.insert_many(default_cats)

    # Seed admin/staff users if not exist
    admin_email = os.environ['ADMIN_EMAIL']
    if not await db.users.find_one({'email': admin_email}):
        await db.users.insert_one({
            'user_id': f"user_{uuid.uuid4().hex[:12]}",
            'email': admin_email,
            'name': 'Retro Farms Admin',
            'phone': '',
            'role': 'admin',
            'password_hash': hash_password(os.environ['ADMIN_PASSWORD']),
            'created_at': now_utc(),
            'provider': 'email',
        })
    staff_email = os.environ['STAFF_EMAIL']
    if not await db.users.find_one({'email': staff_email}):
        await db.users.insert_one({
            'user_id': f"user_{uuid.uuid4().hex[:12]}",
            'email': staff_email,
            'name': 'Retro Farms Staff',
            'phone': '',
            'role': 'staff',
            'password_hash': hash_password(os.environ['STAFF_PASSWORD']),
            'created_at': now_utc(),
            'provider': 'email',
        })

    # Seed farmers if none exist
    if await db.farmers.count_documents({}) == 0:
        seed_farmers = [
            {
                'farmer_id': f"farmer_{uuid.uuid4().hex[:10]}",
                'name': 'Dr. Venkat', 'creds': 'M.Sc, Ph.D in Chemistry', 'role': 'Founder & Farm Director',
                'photo': 'https://customer-assets-39nsmqrw.emergentagent.net/job_farm-to-table-541/artifacts/t0o5gwzu_Dr.Avudoddi.Venkat.jpeg',
                'order': 1, 'created_at': now_utc(),
            },
            {
                'farmer_id': f"farmer_{uuid.uuid4().hex[:10]}",
                'name': 'Mr. Avudoddi Ramakrishna', 'creds': 'MBA', 'role': 'Operations & Distribution',
                'photo': 'https://customer-assets-39nsmqrw.emergentagent.net/job_farm-to-table-541/artifacts/51ekxg06_Avudoddi.Ramakrishna.jpeg',
                'order': 2, 'created_at': now_utc(),
            },
            {
                'farmer_id': f"farmer_{uuid.uuid4().hex[:10]}",
                'name': 'Mr. Avudoddi Mallikarjun', 'creds': 'M.Sc, Ph.D in Chemistry', 'role': 'Livestock & Nutrition Lead',
                'photo': 'https://customer-assets-39nsmqrw.emergentagent.net/job_farm-to-table-541/artifacts/2qghb75g_Avudoddi.Mallikarjun.jpeg',
                'order': 3, 'created_at': now_utc(),
            },
        ]
        await db.farmers.insert_many(seed_farmers)

@app.on_event("shutdown")
async def shutdown():
    client.close()

# ==================== AUTH ROUTES ====================

@api.get("/")
async def root():
    return {"message": "Retro Farms API"}

@api.post("/auth/session")
async def create_session(request: Request, response: Response):
    """Exchange Emergent session_id for our session_token cookie."""
    body = await request.json()
    session_id = body.get('session_id')
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    async with httpx.AsyncClient(timeout=15) as h:
        r = await h.get(
            "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
            headers={'X-Session-ID': session_id},
        )
    if r.status_code != 200:
        raise HTTPException(status_code=401, detail="Invalid session_id")
    data = r.json()
    email = data['email']
    # find or create user
    existing = await db.users.find_one({'email': email})
    if existing:
        user_id = existing['user_id']
        await db.users.update_one(
            {'user_id': user_id},
            {'$set': {'name': data.get('name', existing.get('name', '')),
                       'picture': data.get('picture', existing.get('picture', ''))}},
        )
        role = existing.get('role', 'customer')
    else:
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        role = 'customer'
        await db.users.insert_one({
            'user_id': user_id,
            'email': email,
            'name': data.get('name', ''),
            'picture': data.get('picture', ''),
            'role': role,
            'provider': 'google',
            'created_at': now_utc(),
        })
    token = data['session_token']
    expires_at = now_utc() + timedelta(days=7)
    await db.sessions.insert_one({
        'session_token': token,
        'user_id': user_id,
        'expires_at': expires_at,
        'created_at': now_utc(),
    })
    response.set_cookie(
        key='session_token', value=token, httponly=True, secure=True,
        samesite='none', path='/', max_age=7 * 24 * 3600,
    )
    user = await db.users.find_one({'user_id': user_id}, {'_id': 0})
    return user_public(user)

@api.post("/auth/admin-login")
async def admin_login(payload: AdminLoginPayload, response: Response):
    user = await db.users.find_one({'email': payload.email.lower()})
    if not user or not user.get('password_hash') or not check_password(payload.password, user['password_hash']):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    if user.get('role') not in ('admin', 'staff'):
        raise HTTPException(status_code=403, detail="Not an admin/staff account")
    token = f"sess_{uuid.uuid4().hex}"
    expires_at = now_utc() + timedelta(days=7)
    await db.sessions.insert_one({
        'session_token': token, 'user_id': user['user_id'],
        'expires_at': expires_at, 'created_at': now_utc(),
    })
    response.set_cookie(
        key='session_token', value=token, httponly=True, secure=True,
        samesite='none', path='/', max_age=7 * 24 * 3600,
    )
    return user_public(user)

@api.get("/auth/me")
async def me(request: Request):
    u = await get_current_user(request)
    if not u:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return user_public(u)

@api.post("/auth/logout")
async def logout(request: Request, response: Response):
    token = request.cookies.get('session_token')
    if token:
        await db.sessions.delete_one({'session_token': token})
    response.delete_cookie('session_token', path='/')
    return {"ok": True}

@api.patch("/auth/me")
async def update_me(payload: UpdateMePayload, request: Request):
    user = await require_user(request)
    update = {}
    if payload.name is not None:
        update['name'] = payload.name.strip()
    if payload.phone is not None:
        update['phone'] = payload.phone.strip()
    if payload.email is not None:
        new_email = payload.email.strip().lower()
        if new_email != user['email']:
            existing = await db.users.find_one({'email': new_email})
            if existing and existing.get('user_id') != user['user_id']:
                raise HTTPException(status_code=400, detail="Email already in use")
            update['email'] = new_email
    if payload.new_password:
        if len(payload.new_password) < 6:
            raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
        # For email-authenticated accounts, require current password
        if user.get('provider') == 'email':
            if not payload.current_password:
                raise HTTPException(status_code=400, detail="Current password required")
            fresh = await db.users.find_one({'user_id': user['user_id']})
            if not fresh or not fresh.get('password_hash') or not check_password(payload.current_password, fresh['password_hash']):
                raise HTTPException(status_code=401, detail="Current password is incorrect")
        update['password_hash'] = hash_password(payload.new_password)
        update['provider'] = 'email'
    if update:
        await db.users.update_one({'user_id': user['user_id']}, {'$set': update})
    doc = await db.users.find_one({'user_id': user['user_id']}, {'_id': 0, 'password_hash': 0})
    return user_public(doc)

# ==================== PRODUCT ROUTES ====================

@api.get("/products")
async def list_products():
    docs = await db.products.find({}, {'_id': 0}).to_list(500)
    return docs

@api.get("/products/{slug}")
async def get_product(slug: str):
    p = await db.products.find_one({'slug': slug}, {'_id': 0})
    if not p:
        raise HTTPException(status_code=404, detail="Not found")
    return p

@api.patch("/admin/products/{slug}/variants/{variant_id}/stock")
async def update_variant_stock(slug: str, variant_id: str, payload: UpdateStockPayload, _user=Depends(require_admin_or_staff)):
    p = await db.products.find_one({'slug': slug})
    if not p:
        raise HTTPException(status_code=404, detail="Not found")
    variants = p['variants']
    for v in variants:
        if v['id'] == variant_id:
            v['stock'] = max(0, int(payload.stock))
            break
    await db.products.update_one({'slug': slug}, {'$set': {'variants': variants}})
    return {"ok": True}

# ==================== ORDER ROUTES ====================

def calc_totals(items_resolved):
    subtotal = sum(i['price'] * i['qty'] for i in items_resolved)
    delivery = 100 if subtotal < 200 else 0
    total = subtotal + delivery
    return subtotal, delivery, total

async def resolve_items(items):
    resolved = []
    for it in items:
        p = await db.products.find_one({'slug': it.slug}, {'_id': 0})
        if not p:
            raise HTTPException(status_code=400, detail=f"Product {it.slug} not found")
        variant = next((v for v in p['variants'] if v['id'] == it.variant_id), None)
        if not variant:
            raise HTTPException(status_code=400, detail=f"Variant {it.variant_id} not found")
        entry = {
            'slug': p['slug'], 'name': p['name'], 'image': p['image'],
            'variant_id': variant['id'], 'variant_label': variant['label'],
            'price': variant['price'], 'qty': it.qty,
        }
        opts = getattr(it, 'options', None)
        if opts:
            entry['options'] = opts
        resolved.append(entry)
    return resolved

@api.post("/orders/create")
async def create_order(payload: OrderCreatePayload, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Sign in to place order")
    if not payload.items:
        raise HTTPException(status_code=400, detail="Cart empty")
    items = await resolve_items(payload.items)
    subtotal, delivery, total = calc_totals(items)
    order_id = uuid.uuid4().hex[:8]
    order_doc = {
        'order_id': order_id,
        'user_id': user['user_id'],
        'customer_email': user['email'],
        'customer_name': user.get('name', ''),
        'address': payload.address.dict(),
        'items': items,
        'subtotal': subtotal,
        'delivery_charge': delivery,
        'total': total,
        'payment_method': payload.payment_method,
        'payment_status': 'pending',
        'status': 'Placed',
        'assigned_staff_id': None,
        'assigned_staff_name': None,
        'created_at': now_utc(),
    }

    if payload.payment_method == 'razorpay':
        rzp_order = rzp.order.create({
            'amount': total * 100,  # paise
            'currency': 'INR',
            'receipt': f"rf_{order_id}",
            'payment_capture': 1,
            'notes': {'internal_order_id': order_id, 'email': user['email']},
        })
        order_doc['razorpay_order_id'] = rzp_order['id']
        await db.orders.insert_one(order_doc)
        return {
            'order_id': order_id,
            'razorpay_order_id': rzp_order['id'],
            'amount': total * 100,
            'currency': 'INR',
            'key_id': RZP_KEY_ID,
            'subtotal': subtotal, 'delivery': delivery, 'total': total,
        }
    else:  # cod
        order_doc['payment_status'] = 'Cod Pending'
        await db.orders.insert_one(order_doc)
        # decrement stock
        await decrement_stock(items)
        return {'order_id': order_id, 'total': total, 'payment_method': 'cod', 'subtotal': subtotal, 'delivery': delivery}

async def decrement_stock(items):
    for it in items:
        p = await db.products.find_one({'slug': it['slug']})
        if not p:
            continue
        variants = p['variants']
        for v in variants:
            if v['id'] == it['variant_id']:
                v['stock'] = max(0, v['stock'] - it['qty'])
        await db.products.update_one({'slug': it['slug']}, {'$set': {'variants': variants}})

@api.post("/orders/verify")
async def verify_payment(payload: PaymentVerifyPayload, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Sign in required")
    order = await db.orders.find_one({'order_id': payload.order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    body = f"{payload.razorpay_order_id}|{payload.razorpay_payment_id}"
    expected = hmac.new(RZP_KEY_SECRET.encode(), body.encode(), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(expected, payload.razorpay_signature):
        await db.orders.update_one({'order_id': payload.order_id}, {'$set': {'payment_status': 'Failed'}})
        raise HTTPException(status_code=400, detail="Payment signature invalid")
    await db.orders.update_one(
        {'order_id': payload.order_id},
        {'$set': {
            'payment_status': 'Paid',
            'razorpay_payment_id': payload.razorpay_payment_id,
            'razorpay_signature': payload.razorpay_signature,
            'paid_at': now_utc(),
        }},
    )
    await decrement_stock(order['items'])
    return {"ok": True, "order_id": payload.order_id}

@api.get("/orders/my")
async def my_orders(request: Request):
    user = await require_user(request)
    docs = await db.orders.find({'user_id': user['user_id']}, {'_id': 0}).sort('created_at', -1).to_list(200)
    return docs

@api.get("/orders/{order_id}")
async def get_order(order_id: str, request: Request):
    user = await require_user(request)
    doc = await db.orders.find_one({'order_id': order_id}, {'_id': 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    if user['role'] not in ('admin', 'staff') and doc['user_id'] != user['user_id']:
        raise HTTPException(status_code=403, detail="Forbidden")
    return doc

# ==================== ADMIN ROUTES ====================

@api.get("/admin/stats")
async def admin_stats(_u=Depends(require_admin_or_staff)):
    # Aggregate-based stats — no full-document loads
    total_orders = await db.orders.count_documents({})
    pending_orders = await db.orders.count_documents({'status': {'$nin': ['Delivered', 'Cancelled']}})
    products_count = await db.products.count_documents({})
    customers_count = await db.users.count_documents({'role': 'customer'})
    revenue_pipe = [
        {'$match': {'payment_status': 'Paid'}},
        {'$group': {'_id': None, 'total': {'$sum': '$total'}}},
    ]
    rev_res = await db.orders.aggregate(revenue_pipe).to_list(1)
    revenue = rev_res[0]['total'] if rev_res else 0
    return {
        'revenue': revenue,
        'orders': total_orders,
        'pending': pending_orders,
        'products': products_count,
        'customers': customers_count,
    }

@api.get("/admin/orders")
async def admin_orders(limit: int = 500, skip: int = 0, _u=Depends(require_admin_or_staff)):
    docs = await db.orders.find({}, {'_id': 0}).sort('created_at', -1).skip(skip).limit(limit).to_list(limit)
    return docs

@api.patch("/admin/orders/{order_id}")
async def admin_update_order(order_id: str, payload: UpdateOrderPayload, _u=Depends(require_admin_or_staff)):
    order = await db.orders.find_one({'order_id': order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Not found")
    update = {}
    if payload.status:
        update['status'] = payload.status
        if payload.status == 'Delivered':
            update['delivered_at'] = now_utc()
            if order.get('payment_method') == 'cod':
                update['payment_status'] = 'Paid'
    if payload.assigned_staff_id is not None:
        if payload.assigned_staff_id == '':
            update['assigned_staff_id'] = None
            update['assigned_staff_name'] = None
        else:
            staff = await db.users.find_one({'user_id': payload.assigned_staff_id})
            if not staff:
                raise HTTPException(status_code=400, detail="Staff not found")
            update['assigned_staff_id'] = staff['user_id']
            update['assigned_staff_name'] = staff['name']
    if update:
        await db.orders.update_one({'order_id': order_id}, {'$set': update})
    doc = await db.orders.find_one({'order_id': order_id}, {'_id': 0})
    return doc

@api.get("/admin/customers")
async def admin_customers(limit: int = 500, _u=Depends(require_admin_or_staff)):
    # Aggregate order stats per user (in-database)
    pipe = [
        {'$match': {'status': {'$ne': 'Cancelled'}}},
        {'$group': {'_id': '$user_id',
                    'total_spent': {'$sum': '$total'},
                    'orders': {'$sum': 1}}},
    ]
    stats = {r['_id']: r async for r in db.orders.aggregate(pipe)}
    # Also count total (including cancelled) for orders shown
    total_pipe = [{'$group': {'_id': '$user_id', 'orders': {'$sum': 1}}}]
    totals = {r['_id']: r['orders'] async for r in db.orders.aggregate(total_pipe)}

    users = await db.users.find({'role': 'customer'},
                                {'_id': 0, 'user_id': 1, 'name': 1, 'email': 1, 'phone': 1}
                                ).limit(limit).to_list(limit)
    out = []
    for u in users:
        s = stats.get(u['user_id'], {})
        out.append({
            'user_id': u['user_id'],
            'name': u.get('name', ''),
            'email': u['email'],
            'phone': u.get('phone', '') or '—',
            'orders': totals.get(u['user_id'], 0),
            'total_spent': s.get('total_spent', 0),
        })
    out.sort(key=lambda x: x['total_spent'], reverse=True)
    return out

@api.get("/admin/customers/{user_id}/orders")
async def admin_customer_orders(user_id: str, limit: int = 200, _u=Depends(require_admin_or_staff)):
    docs = await db.orders.find({'user_id': user_id}, {'_id': 0}).sort('created_at', -1).limit(limit).to_list(limit)
    user = await db.users.find_one({'user_id': user_id},
                                    {'_id': 0, 'user_id': 1, 'name': 1, 'email': 1, 'phone': 1, 'provider': 1})
    return {'user': user, 'orders': docs}

@api.get("/admin/staff")
async def admin_staff_list(_u=Depends(require_admin_or_staff)):
    users = await db.users.find({'role': {'$in': ['admin', 'staff']}}, {'_id': 0, 'password_hash': 0}).to_list(500)
    return [{'user_id': u['user_id'], 'name': u['name'], 'email': u['email'],
             'role': u['role'].capitalize(), 'phone': u.get('phone', '') or '—'} for u in users]

@api.post("/admin/staff")
async def admin_staff_create(payload: StaffCreatePayload, _u=Depends(require_admin)):
    email = payload.email.lower()
    if await db.users.find_one({'email': email}):
        raise HTTPException(status_code=400, detail="Email already exists")
    if payload.role not in ('staff', 'admin'):
        raise HTTPException(status_code=400, detail="Invalid role")
    doc = {
        'user_id': f"user_{uuid.uuid4().hex[:12]}",
        'email': email,
        'name': payload.name,
        'phone': payload.phone or '',
        'role': payload.role,
        'password_hash': hash_password(payload.password),
        'created_at': now_utc(),
        'provider': 'email',
    }
    await db.users.insert_one(doc)
    return {'user_id': doc['user_id'], 'name': doc['name'], 'email': doc['email'],
            'role': doc['role'].capitalize(), 'phone': doc['phone'] or '—'}

@api.delete("/admin/staff/{user_id}")
async def admin_staff_delete(user_id: str, current=Depends(require_admin)):
    if user_id == current['user_id']:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    r = await db.users.delete_one({'user_id': user_id, 'role': {'$in': ['admin', 'staff']}})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True}

# ==================== OFFLINE ORDERS ====================

@api.post("/admin/orders/offline")
async def create_offline_order(payload: OfflineOrderPayload, current=Depends(require_admin_or_staff)):
    if not payload.items:
        raise HTTPException(status_code=400, detail="Items required")
    items = await resolve_items(payload.items)
    subtotal, delivery, total = calc_totals(items)
    if payload.total_override is not None and payload.total_override >= 0:
        total = int(payload.total_override)
        delivery = max(0, total - subtotal)

    # Find or create pseudo customer user — prefer by phone (unique identifier for offline)
    email = (payload.customer_email or '').lower().strip()
    phone = (payload.customer_phone or '').strip()
    existing = None
    if phone:
        existing = await db.users.find_one({'phone': phone, 'role': 'customer'})
    if not existing and email:
        existing = await db.users.find_one({'email': email})
    if not email:
        email = f"offline_{phone or uuid.uuid4().hex[:8]}@retrofarms.offline"

    address_to_save = payload.address or None

    if existing:
        cust_id = existing['user_id']
        update = {
            'name': payload.customer_name or existing.get('name', ''),
            'phone': phone or existing.get('phone', ''),
        }
        if email and email != existing.get('email'):
            # only update email if it's a real one, not a synthetic offline
            if not existing.get('email', '').endswith('@retrofarms.offline'):
                pass  # keep original real email
            else:
                update['email'] = email
        if payload.save_customer_address and address_to_save:
            update['saved_address'] = address_to_save
        await db.users.update_one({'user_id': cust_id}, {'$set': update})
    else:
        cust_id = f"user_{uuid.uuid4().hex[:12]}"
        doc = {
            'user_id': cust_id,
            'email': email,
            'name': payload.customer_name,
            'phone': phone,
            'role': 'customer',
            'provider': 'offline',
            'created_at': now_utc(),
        }
        if payload.save_customer_address and address_to_save:
            doc['saved_address'] = address_to_save
        await db.users.insert_one(doc)

    order_id = uuid.uuid4().hex[:8]
    order_doc = {
        'order_id': order_id,
        'user_id': cust_id,
        'customer_email': email,
        'customer_name': payload.customer_name,
        'address': payload.address or {'full_name': payload.customer_name, 'phone': phone,
                                        'line1': 'Offline order', 'city': 'Hyderabad', 'pincode': '', 'landmark': ''},
        'items': items,
        'subtotal': subtotal,
        'delivery_charge': delivery,
        'total': total,
        'payment_method': payload.payment_method,
        'payment_status': payload.payment_status,
        'status': payload.status,
        'assigned_staff_id': None,
        'assigned_staff_name': None,
        'source': 'offline',
        'notes': payload.notes or '',
        'created_at': now_utc(),
        'created_by': current['user_id'],
    }
    await db.orders.insert_one(order_doc)
    await decrement_stock(items)
    doc = await db.orders.find_one({'order_id': order_id}, {'_id': 0})
    return doc

# Lookup a customer by phone (for offline order autofill)
@api.get("/admin/customers/lookup")
async def lookup_customer(phone: str = Query(...), _u=Depends(require_admin_or_staff)):
    u = await db.users.find_one({'phone': phone, 'role': 'customer'},
                                 {'_id': 0, 'user_id': 1, 'name': 1, 'email': 1, 'phone': 1, 'saved_address': 1})
    return u or {}

# Update a customer's profile fields (no impact on past orders)
@api.patch("/admin/customers/{user_id}")
async def update_customer(user_id: str, payload: CustomerUpdatePayload, _u=Depends(require_admin_or_staff)):
    update = {}
    if payload.name is not None: update['name'] = payload.name.strip()
    if payload.phone is not None: update['phone'] = payload.phone.strip()
    if payload.email is not None:
        new_email = payload.email.strip().lower()
        if new_email:
            existing = await db.users.find_one({'email': new_email})
            if existing and existing.get('user_id') != user_id:
                raise HTTPException(status_code=400, detail="Email already in use")
            update['email'] = new_email
    if payload.address is not None:
        update['saved_address'] = payload.address
    if update:
        r = await db.users.update_one({'user_id': user_id, 'role': 'customer'}, {'$set': update})
        if r.matched_count == 0:
            raise HTTPException(status_code=404, detail="Customer not found")
    doc = await db.users.find_one({'user_id': user_id},
                                    {'_id': 0, 'user_id': 1, 'name': 1, 'email': 1, 'phone': 1, 'saved_address': 1})
    return doc

# ==================== PRODUCT CRUD ====================

@api.post("/admin/products")
async def create_product(payload: ProductCreatePayload, _u=Depends(require_admin)):
    slug = payload.slug.strip().lower().replace(' ', '-')
    if await db.products.find_one({'slug': slug}):
        raise HTTPException(status_code=400, detail="Slug already exists")
    doc = {
        'slug': slug,
        'name': payload.name,
        'category': payload.category,
        'image': payload.image,
        'from_price': payload.from_price,
        'description': payload.description,
        'variants': [v.dict() for v in payload.variants],
    }
    await db.products.insert_one(doc)
    doc.pop('_id', None)
    return doc

@api.put("/admin/products/{slug}")
async def update_product(slug: str, payload: ProductUpdatePayload, _u=Depends(require_admin)):
    p = await db.products.find_one({'slug': slug})
    if not p:
        raise HTTPException(status_code=404, detail="Not found")
    update = {k: v for k, v in payload.dict().items() if v is not None}
    if update:
        await db.products.update_one({'slug': slug}, {'$set': update})
    doc = await db.products.find_one({'slug': slug}, {'_id': 0})
    return doc

@api.delete("/admin/products/{slug}")
async def delete_product(slug: str, _u=Depends(require_admin)):
    r = await db.products.delete_one({'slug': slug})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True}

@api.post("/admin/products/{slug}/variants")
async def add_variant(slug: str, variant: VariantInput, _u=Depends(require_admin)):
    p = await db.products.find_one({'slug': slug})
    if not p:
        raise HTTPException(status_code=404, detail="Not found")
    if any(v['id'] == variant.id for v in p['variants']):
        raise HTTPException(status_code=400, detail="Variant id already exists")
    p['variants'].append(variant.dict())
    await db.products.update_one({'slug': slug}, {'$set': {'variants': p['variants']}})
    doc = await db.products.find_one({'slug': slug}, {'_id': 0})
    return doc

@api.patch("/admin/products/{slug}/variants/{variant_id}")
async def update_variant(slug: str, variant_id: str, payload: VariantUpdatePayload, _u=Depends(require_admin)):
    p = await db.products.find_one({'slug': slug})
    if not p:
        raise HTTPException(status_code=404, detail="Not found")
    update = {k: v for k, v in payload.dict().items() if v is not None}
    if not update:
        return await db.products.find_one({'slug': slug}, {'_id': 0})
    for v in p['variants']:
        if v['id'] == variant_id:
            v.update(update)
            break
    else:
        raise HTTPException(status_code=404, detail="Variant not found")
    await db.products.update_one({'slug': slug}, {'$set': {'variants': p['variants']}})
    doc = await db.products.find_one({'slug': slug}, {'_id': 0})
    return doc

@api.delete("/admin/products/{slug}/variants/{variant_id}")
async def delete_variant(slug: str, variant_id: str, _u=Depends(require_admin)):
    p = await db.products.find_one({'slug': slug})
    if not p:
        raise HTTPException(status_code=404, detail="Not found")
    new_variants = [v for v in p['variants'] if v['id'] != variant_id]
    if len(new_variants) == len(p['variants']):
        raise HTTPException(status_code=404, detail="Variant not found")
    await db.products.update_one({'slug': slug}, {'$set': {'variants': new_variants}})
    return {"ok": True}

# ==================== FARMERS ====================

@api.get("/farmers")
async def list_farmers():
    docs = await db.farmers.find({}, {'_id': 0}).sort('order', 1).to_list(50)
    return docs

@api.post("/admin/farmers")
async def create_farmer(payload: FarmerCreatePayload, _u=Depends(require_admin)):
    doc = {
        'farmer_id': f"farmer_{uuid.uuid4().hex[:10]}",
        'name': payload.name, 'creds': payload.creds, 'role': payload.role,
        'photo': payload.photo, 'order': payload.order, 'created_at': now_utc(),
    }
    await db.farmers.insert_one(doc)
    return {k: v for k, v in doc.items() if k != '_id'}

@api.put("/admin/farmers/{farmer_id}")
async def update_farmer(farmer_id: str, payload: FarmerUpdatePayload, _u=Depends(require_admin)):
    update = {k: v for k, v in payload.dict().items() if v is not None}
    r = await db.farmers.update_one({'farmer_id': farmer_id}, {'$set': update})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="Farmer not found")
    doc = await db.farmers.find_one({'farmer_id': farmer_id}, {'_id': 0})
    return doc

@api.delete("/admin/farmers/{farmer_id}")
async def delete_farmer(farmer_id: str, _u=Depends(require_admin)):
    r = await db.farmers.delete_one({'farmer_id': farmer_id})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True}

# ==================== OFFLINE CUSTOMERS ====================

@api.get("/admin/offline-customers")
async def offline_customers(limit: int = 500, _u=Depends(require_admin_or_staff)):
    """Returns customers with recent order metadata — for quick reuse in offline order form."""
    pipe = [
        {'$sort': {'created_at': -1}},
        {'$group': {
            '_id': '$user_id',
            'orders': {'$sum': 1},
            'last_ordered_at': {'$first': '$created_at'},
            'last_address': {'$first': '$address'},
        }},
    ]
    stats = {r['_id']: r async for r in db.orders.aggregate(pipe)}

    users = await db.users.find({'role': 'customer'},
                                {'_id': 0, 'user_id': 1, 'name': 1, 'email': 1,
                                 'phone': 1, 'provider': 1}
                                ).limit(limit).to_list(limit)
    from datetime import datetime as dt_min
    out = []
    for u in users:
        s = stats.get(u['user_id'], {})
        out.append({
            'user_id': u['user_id'],
            'name': u.get('name', ''),
            'email': u['email'],
            'phone': u.get('phone', ''),
            'provider': u.get('provider', ''),
            'orders': s.get('orders', 0),
            'last_ordered_at': s.get('last_ordered_at'),
            'last_address': s.get('last_address'),
        })
    out.sort(key=lambda x: (x['last_ordered_at'] or dt_min(1970, 1, 1)), reverse=True)
    return out

# ==================== CATEGORIES ====================

@api.get("/categories")
async def list_categories():
    docs = await db.categories.find({}, {'_id': 0}).sort('order', 1).to_list(100)
    return docs

@api.post("/admin/categories")
async def create_category(payload: CategoryPayload, _u=Depends(require_admin)):
    cid = payload.id.strip().lower().replace(' ', '-')
    if not cid or not payload.label:
        raise HTTPException(status_code=400, detail="id and label required")
    if await db.categories.find_one({'id': cid}):
        raise HTTPException(status_code=400, detail="Category id already exists")
    max_order = 0
    async for c in db.categories.find({}, {'order': 1}).sort('order', -1).limit(1):
        max_order = c.get('order', 0)
    doc = {'id': cid, 'label': payload.label.strip(), 'order': max_order + 1}
    await db.categories.insert_one(doc)
    doc.pop('_id', None)
    return doc

@api.patch("/admin/categories/{cid}")
async def update_category(cid: str, payload: CategoryUpdatePayload, _u=Depends(require_admin)):
    update = {k: v for k, v in payload.dict().items() if v is not None}
    r = await db.categories.update_one({'id': cid}, {'$set': update})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    doc = await db.categories.find_one({'id': cid}, {'_id': 0})
    return doc

@api.delete("/admin/categories/{cid}")
async def delete_category(cid: str, reassign_to: Optional[str] = Query(None), _u=Depends(require_admin)):
    in_use = await db.products.count_documents({'category': cid})
    if in_use > 0 and not reassign_to:
        raise HTTPException(status_code=400,
                            detail=f"{in_use} product(s) still use this category. Reassign first via ?reassign_to=<other_id>")
    if reassign_to:
        if not await db.categories.find_one({'id': reassign_to}):
            raise HTTPException(status_code=400, detail="Target category not found")
        await db.products.update_many({'category': cid}, {'$set': {'category': reassign_to}})
    r = await db.categories.delete_one({'id': cid})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True, "reassigned": in_use if reassign_to else 0}

# ==================== REVENUE BREAKDOWN ====================

@api.get("/admin/revenue/breakdown")
async def revenue_breakdown(view: str = Query('day', regex='^(day|week|month|year)$'),
                             start: Optional[str] = None, end: Optional[str] = None,
                             _u=Depends(require_admin_or_staff)):
    """Returns revenue grouped by day/week/month/year for paid orders."""
    match = {'payment_status': 'Paid'}
    if start or end:
        rng = {}
        if start:
            try: rng['$gte'] = datetime.fromisoformat(start)
            except ValueError: pass
        if end:
            try: rng['$lte'] = datetime.fromisoformat(end)
            except ValueError: pass
        if rng: match['created_at'] = rng

    fmt_map = {
        'day': '%Y-%m-%d',
        'week': '%G-W%V',   # ISO week
        'month': '%Y-%m',
        'year': '%Y',
    }
    date_fmt = fmt_map[view]
    pipe = [
        {'$match': match},
        {'$group': {
            '_id': {'$dateToString': {'format': date_fmt, 'date': '$created_at'}},
            'revenue': {'$sum': '$total'},
            'orders': {'$sum': 1},
        }},
        {'$sort': {'_id': 1}},
    ]
    rows = []
    async for r in db.orders.aggregate(pipe):
        avg = round(r['revenue'] / r['orders'], 2) if r['orders'] else 0
        rows.append({'period': r['_id'], 'revenue': r['revenue'], 'orders': r['orders'], 'aov': avg})
    total_rev = sum(r['revenue'] for r in rows)
    total_orders = sum(r['orders'] for r in rows)
    return {
        'view': view,
        'rows': rows,
        'summary': {
            'total_revenue': total_rev,
            'total_orders': total_orders,
            'aov': round(total_rev / total_orders, 2) if total_orders else 0,
        },
    }

# ==================== EXCEL EXPORT ====================

def _flatten_options(opts):
    if not opts: return ''
    if isinstance(opts, dict):
        return ' | '.join(f"{k}: {v}" for k, v in opts.items() if v)
    return str(opts)

@api.get("/admin/orders/export.xlsx")
async def export_orders_xlsx(start: Optional[str] = None, end: Optional[str] = None,
                              status: Optional[str] = None, _u=Depends(require_admin_or_staff)):
    query = {}
    if start or end:
        rng = {}
        if start:
            try: rng['$gte'] = datetime.fromisoformat(start)
            except ValueError: pass
        if end:
            try: rng['$lte'] = datetime.fromisoformat(end)
            except ValueError: pass
        if rng: query['created_at'] = rng
    if status: query['status'] = status

    orders = await db.orders.find(query, {'_id': 0}).sort('created_at', -1).to_list(20000)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Orders'
    headers = ['Order ID', 'Date', 'Customer Name', 'Phone', 'Email', 'Address',
               'Items', 'Chicken Details', 'Payment Method', 'Payment Status',
               'Subtotal', 'Delivery', 'Total', 'Status', 'Assigned To', 'Source', 'Notes']
    ws.append(headers)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2B1D11')
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')

    for o in orders:
        addr = o.get('address') or {}
        addr_str = ', '.join([str(v) for v in [addr.get('line1'), addr.get('line2'), addr.get('city'),
                              addr.get('pincode'), addr.get('landmark')] if v])
        items_str = '; '.join([f"{i.get('name')} ({i.get('variant_label')}) × {i.get('qty')}" for i in o.get('items', [])])
        chicken_details = []
        for i in o.get('items', []):
            opts = i.get('options')
            if opts:
                chicken_details.append(f"{i.get('name')}: {_flatten_options(opts)}")
        created = o.get('created_at')
        if isinstance(created, str):
            try: created = datetime.fromisoformat(created.replace('Z', '+00:00'))
            except Exception: created = None
        row = [
            o.get('order_id', ''),
            created if isinstance(created, datetime) else '',
            o.get('customer_name') or addr.get('full_name', ''),
            addr.get('phone') or '',
            o.get('customer_email', ''),
            addr_str,
            items_str,
            ' | '.join(chicken_details),
            (o.get('payment_method') or '').upper(),
            o.get('payment_status', ''),
            o.get('subtotal', 0),
            o.get('delivery_charge', 0),
            o.get('total', 0),
            o.get('status', ''),
            o.get('assigned_staff_name') or 'Unassigned',
            o.get('source') or 'online',
            o.get('notes', ''),
        ]
        ws.append(row)

    # Format date column
    for cell in ws['B'][1:]:
        cell.number_format = 'yyyy-mm-dd hh:mm'
    for col in ['K', 'L', 'M']:  # money columns
        for cell in ws[col][1:]:
            cell.number_format = '"₹"#,##0'

    # Sensible widths
    widths = [12, 20, 22, 14, 26, 40, 50, 30, 14, 14, 10, 10, 10, 14, 18, 10, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"retrofarms_orders_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )

# ==================== FEEDBACK ====================

FEEDBACK_EDIT_WINDOW_HOURS = 48
FEEDBACK_RATE_LIMIT_PER_HOUR = 5
FEEDBACK_INSTANT_THRESHOLD_SEC = 5

def _client_ip(request: Request) -> str:
    xff = request.headers.get('x-forwarded-for')
    if xff:
        return xff.split(',')[0].strip()
    return request.client.host if request.client else ''

@api.post("/feedback")
async def submit_feedback(payload: FeedbackCreatePayload, request: Request):
    user = await require_user(request)
    # Block admins/staff from leaving feedback
    if user.get('role') in ('admin', 'staff'):
        raise HTTPException(status_code=403, detail="Staff and admins cannot submit customer feedback")

    order = await db.orders.find_one({'order_id': payload.order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.get('user_id') != user['user_id']:
        raise HTTPException(status_code=403, detail="You can only review your own orders")
    if order.get('status') != 'Delivered':
        raise HTTPException(status_code=400, detail="Feedback is only available after your order is delivered")

    if await db.feedback.find_one({'order_id': payload.order_id}):
        raise HTTPException(status_code=400, detail="Feedback already submitted for this order — you can edit it within 48 hours")

    ip = _client_ip(request)
    now = now_utc()
    # Rate limit: max 5 feedbacks per user OR ip in the last hour
    one_hour_ago = now - timedelta(hours=1)
    recent_by_user = await db.feedback.count_documents({'user_id': user['user_id'], 'created_at': {'$gte': one_hour_ago}})
    recent_by_ip = await db.feedback.count_documents({'ip_address': ip, 'created_at': {'$gte': one_hour_ago}}) if ip else 0
    if recent_by_user >= FEEDBACK_RATE_LIMIT_PER_HOUR or recent_by_ip >= FEEDBACK_RATE_LIMIT_PER_HOUR:
        raise HTTPException(status_code=429, detail="Too many feedback submissions. Please try again later.")

    # Anti-fraud heuristics — mark 'flagged' for admin review instead of blocking
    flags = []
    delivered_at = order.get('delivered_at')
    # Guess delivered_at from status change: use created_at diff if we don't track it — fallback to now
    if delivered_at:
        elapsed = (now - (delivered_at if delivered_at.tzinfo else delivered_at.replace(tzinfo=timezone.utc))).total_seconds()
        if elapsed < FEEDBACK_INSTANT_THRESHOLD_SEC:
            flags.append('instant_after_delivery')
    comment = (payload.comment or '').strip()
    if comment:
        dup = await db.feedback.find_one({'comment': comment, 'user_id': {'$ne': user['user_id']}})
        if dup:
            flags.append('duplicate_comment')
    # If IP has multiple recent submissions
    if recent_by_ip >= 3:
        flags.append('high_ip_volume')
    if recent_by_user >= 3:
        flags.append('high_user_volume')

    fb_id = f"fb_{uuid.uuid4().hex[:12]}"
    doc = {
        'feedback_id': fb_id,
        'order_id': payload.order_id,
        'user_id': user['user_id'],
        'customer_email': user.get('email', ''),
        'customer_name': user.get('name', ''),
        'customer_phone': user.get('phone', '') or (order.get('address') or {}).get('phone', ''),
        'rating': int(payload.rating),
        'comment': comment,
        'ip_address': ip,
        'user_agent': request.headers.get('user-agent', '')[:400],
        'status': 'flagged' if flags else 'active',
        'flags': flags,
        'admin_response': '',
        'edit_count': 0,
        'created_at': now,
        'updated_at': now,
    }
    try:
        await db.feedback.insert_one(doc)
    except Exception:
        raise HTTPException(status_code=400, detail="Feedback already exists for this order")
    doc.pop('_id', None)
    return doc

@api.patch("/feedback/{feedback_id}")
async def edit_feedback(feedback_id: str, payload: FeedbackEditPayload, request: Request):
    user = await require_user(request)
    fb = await db.feedback.find_one({'feedback_id': feedback_id})
    if not fb:
        raise HTTPException(status_code=404, detail="Feedback not found")
    if fb['user_id'] != user['user_id']:
        raise HTTPException(status_code=403, detail="You can only edit your own feedback")
    created = fb['created_at']
    if isinstance(created, str):
        created = datetime.fromisoformat(created)
    if created.tzinfo is None:
        created = created.replace(tzinfo=timezone.utc)
    if now_utc() - created > timedelta(hours=FEEDBACK_EDIT_WINDOW_HOURS):
        raise HTTPException(status_code=400, detail=f"The {FEEDBACK_EDIT_WINDOW_HOURS}-hour edit window has closed")
    update = {'updated_at': now_utc(), 'edit_count': fb.get('edit_count', 0) + 1}
    if payload.rating is not None:
        update['rating'] = int(payload.rating)
    if payload.comment is not None:
        update['comment'] = payload.comment.strip()
    await db.feedback.update_one({'feedback_id': feedback_id}, {'$set': update})
    doc = await db.feedback.find_one({'feedback_id': feedback_id}, {'_id': 0})
    return doc

@api.get("/feedback/order/{order_id}")
async def get_order_feedback(order_id: str, request: Request):
    user = await get_current_user(request)
    fb = await db.feedback.find_one({'order_id': order_id}, {'_id': 0})
    if not fb:
        return None
    # Only owner or admin/staff can see full details; public callers get sanitized (or nothing)
    if not user:
        return None
    if user.get('role') not in ('admin', 'staff') and fb['user_id'] != user['user_id']:
        return None
    return fb

@api.get("/admin/feedback")
async def admin_list_feedback(status: Optional[str] = None, limit: int = 500,
                               _u=Depends(require_admin_or_staff)):
    q = {}
    if status: q['status'] = status
    docs = await db.feedback.find(q, {'_id': 0}).sort('created_at', -1).limit(limit).to_list(limit)
    return docs

@api.patch("/admin/feedback/{feedback_id}")
async def moderate_feedback(feedback_id: str, payload: FeedbackModeratePayload,
                             current=Depends(require_admin_or_staff)):
    update = {}
    if payload.status:
        if payload.status not in ('active', 'flagged', 'hidden'):
            raise HTTPException(status_code=400, detail="Invalid status")
        update['status'] = payload.status
        update['moderated_by'] = current['user_id']
        update['moderated_at'] = now_utc()
    if payload.admin_response is not None:
        update['admin_response'] = payload.admin_response.strip()
    if update:
        r = await db.feedback.update_one({'feedback_id': feedback_id}, {'$set': update})
        if r.matched_count == 0:
            raise HTTPException(status_code=404, detail="Not found")
    doc = await db.feedback.find_one({'feedback_id': feedback_id}, {'_id': 0})
    return doc

@api.get("/feedback/public")
async def public_feedback(limit: int = 20):
    """Only active (moderated-safe) feedback for storefront display."""
    docs = await db.feedback.find({'status': 'active', 'rating': {'$gte': 4}},
                                   {'_id': 0, 'ip_address': 0, 'user_agent': 0, 'customer_email': 0, 'customer_phone': 0}
                                   ).sort('created_at', -1).limit(limit).to_list(limit)
    return docs

# ==================== APP SETUP ====================

app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
